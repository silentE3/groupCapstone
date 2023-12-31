'''config holds the logic to read in a configuration object'''
import json
from io import TextIOWrapper, StringIO
import logging
from openpyxl import load_workbook
from app import models
from app.models import Configuration, NoSurveyGroupMethodConsts

__logger = logging.getLogger(__name__)

def read_json(config_path: str) -> Configuration:
    """Reads in a json configuration file"""
    with open(config_path, mode="r", encoding="UTF-8") as json_file:
        data: Configuration = read_json_from_io(json_file)


# The following global module variables and assignments are a move towards
# creating a single instance of the configuration that does not need to be
# passed to each method.
    # pylint: disable=global-statement
    global CONFIG_DATA
    CONFIG_DATA = data

    return data


def read_json_from_io(text_buffer: TextIOWrapper) -> Configuration:
    '''
    Reads in json configuration data from an io buffer version of the data.
    '''
    text_buffer.seek(0)
    data: Configuration = json.load(text_buffer)
    __check_config_validity(data)

    try:
        data["prioritize_preferred_over_availability"]
    except KeyError:
        data["prioritize_preferred_over_availability"] = False

    # pylint: disable=global-statement
    global CONFIG_DATA
    CONFIG_DATA = data

    return data


def read_report_config(report_filename: str) -> Configuration:
    '''
    Reads in json configuration data from the 'config' tab of an existing xlsx report file.
    '''
    config_sheet_name: str = "config"

    # load the workbook
    report_workbook = load_workbook(report_filename)
    if config_sheet_name not in report_workbook.sheetnames:
        raise KeyError(
            "Unable to load the config data from the report file. 'config' tab does not exist.")
    # get the workbook's config sheet
    config_sheet = report_workbook['config']

    # Initialize empty dictionaries for the necessary config elements
    config_data = {}
    field_mappings = {}
    report_fields = {}

    # iterate through the config sheet by column, since each config entry was stored as a separate column
    for col in config_sheet.iter_cols():
        config_item_value = []
        config_item_key: str = ""
        for row_num, cell in enumerate(col):
            if row_num == 0:
                # The header of the column (first row) contains the item's key
                config_item_key = str(cell.value)
            elif cell.value is None:
                continue
            else:
                config_item_value.append(cell.value)

        if len(config_item_value) == 1:
            # if there is only one value for the config item, we don't want to store it in a list
            config_item_value = config_item_value[0]

        # field names/mappings are stored in an internal dictionary
        if "field_name" in config_item_key:
            field_mappings[config_item_key] = config_item_value
        # report fields (which contain "show_") are stored in an internal dictionary
        # NOTE: currently only report fields contain "show_". This will need to be revisited if
        #       that changes.
        elif "show_" in config_item_key:
            report_fields[config_item_key] = config_item_value
        else:
            config_data[config_item_key] = config_item_value

    config_data["field_mappings"] = field_mappings
    config_data["report_fields"] = report_fields

    # convert the dict to json data and write it to an io bufffer
    text_buffer = StringIO()
    text_buffer.write(json.dumps(config_data))

    # load/read the config data from the io buffer and return it
    data: Configuration = read_json_from_io(text_buffer)

    # pylint: disable=global-statement
    global CONFIG_DATA
    CONFIG_DATA = data

    return data


def __check_config_validity(config_data: Configuration):

    '''
    __check_config_validity() helps determin how students who did not fill out a survey will be distributed across groups, STANDARD is
    just the same way it was being hadles, no particular attention paid to the students with no survey data,
    DISTRIBUTE_EVENLY will distribute the students with no survey data evenly across the groups, GROUP_TOGETHER will
    put all the students with no survey data in the same group
    '''

    valid_no_survey_group_methods = [NoSurveyGroupMethodConsts.STANDARD_GROUPING,
                                     NoSurveyGroupMethodConsts.DISTRIBUTE_EVENLY, NoSurveyGroupMethodConsts.GROUP_TOGETHER]
    if "no_survey_group_method" not in config_data:
        config_data["no_survey_group_method"] = NoSurveyGroupMethodConsts.STANDARD_GROUPING
    if config_data['no_survey_group_method'] not in valid_no_survey_group_methods:
        __logger.error('Invalid configuration selection for "no_survey_group_method".')
        raise ValueError('Invalid configuration selection for "no_survey_group_method".')

def validate_field_mappings(fields: models.SurveyFieldMapping):
    '''
    Validates the field mappings specification in the configuration data.
    '''
    valid_fields = True
    if fields.get('availability_field_names') is None or len(fields.get('availability_field_names')) == 0:
        __logger.error(__field_error_msg('availability_field_names'))
        valid_fields = False

    if fields.get('disliked_students_field_names') is None or len(fields.get('disliked_students_field_names')) == 0:
        __logger.error(__field_error_msg('disliked_students_field_names'))
        valid_fields = False

    if fields.get('preferred_students_field_names') is None or len(fields.get('preferred_students_field_names')) == 0:
        __logger.error(__field_error_msg('preferred_students_field_names'))
        valid_fields = False

    if fields.get('student_id_field_name') is None:
        __logger.error(__field_error_msg('student_id_field_name'))
        valid_fields = False

    if valid_fields is False:
        raise AttributeError('Invalid or missing field mappings in the configuration file.')


def __field_error_msg(field_name: str) -> str:
    return f'Error: No {field_name} field name was specified in the configuration file. Please provide a value for "{field_name}".'


CONFIG_DATA = None
