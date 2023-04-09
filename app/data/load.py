"""
Provides the ability to load data into the program including raw survey data and grouped data
"""

import copy
import csv
from io import TextIOWrapper
from io import StringIO
import logging
import re
import datetime as dt
from typing import Union
from openpyxl import load_workbook
from app import models
from app.group import validate
from app import config

__logger = logging.getLogger(__name__)


def parse_asurite(val: str) -> str:
    '''
    parses a student's id from the string.
    Returns the first value when splitting a str on a space character
    '''
    return re.search(r'\S+', val).group()


def split_on_delimiters(availability: str, delimiters: str):
    '''
    allows handling of as many delimiters as the user wants to define in the config file, used by parse_survey_record()
    '''
    if len(delimiters) == 0:
        raise ValueError(
            "Configuration file has no availability delimiters defined")
    delim_chars = f'[{re.escape(delimiters)}]'

    return re.split(delim_chars, availability)


def total_availability_matches(student: models.SurveyRecord, students: list[models.SurveyRecord]) -> int:
    """
    checks the student's availability against everyone elses and counts the number of times that they match
    Args:
        student (models.SurveyRecord): student to check matches for
        students (list[models.SurveyRecord]): full list of students
    Returns:
        int: number of times the student matches availability in the dataset
    """

    totals = 0

    for other_student in students:
        if student.student_id == other_student.student_id:
            continue

        for time, avail_days in other_student.availability.items():
            for avail_day in avail_days:
                if avail_day in student.availability[time] and not avail_day == '':
                    totals += 1

    return totals


def wildcard_availability(availability_fields: list) -> dict[str, list[str]]:
    '''
    constructs availability to match all of the time fields provided for any given day
    '''
    avail = {}
    for field in availability_fields:
        avail[field] = validate.WEEK_DAYS

    return avail


def has_availability(student: models.SurveyRecord) -> bool:
    '''
    checks if the student marked any days that they were available during the allotted time
    '''
    avail = 0
    for days in student.availability.values():
        avail += len(days)

    return avail > 0


def preprocess_survey_data(students: list[models.SurveyRecord], field_mapping: models.SurveyFieldMapping):
    '''
    performs some basic "preprocessing" of the survey data to ensure the grouping algorithm can function expected.
    This includes the following:
    - checking for students that didn't add availability and setting it to match any
    - checking for students that have availability that didn't match to anyone else's
    - checking for students that have preferred students that in turn disliked them (Not yet implemented)
    - checking for students that have preferred students that didn't match in their availability (Not yet implemented)
    - checking for students that put themselves as preferred or disliked and removing them from the preferred/disliked lists
    - checking for students that could "possibly" be paired with one of their preferred students (they provided
        preferred student(s) and at list one of these students didn't list them as "disliked")
    '''
    for student in students:
        if not student.provided_availability:
            print(f"student '{student.student_id}' did not provide any availability")
            student.availability = wildcard_availability(
                field_mapping["availability_field_names"])
        if total_availability_matches(student, students) == 0:
            print(f"student '{student.student_id}' did not have matching availability with anyone else")
            student.has_matching_availability = False
        # remove self from list of disliked
        if student.student_id in student.disliked_students:
            student.disliked_students.remove(student.student_id)
        # remove self from list of preferred
        if student.student_id in student.preferred_students:
            student.preferred_students.remove(student.student_id)
        # did student provide preferred student selection
        student.provided_pref_students = len(student.disliked_students) > 0
        # could student "possibly" (reasonably) be paired with one of their preferred selections
        student.pref_pairing_possible = validate.student_pref_pair_possible(
            students, student)


def parse_survey_record(field_mapping: models.SurveyFieldMapping, row: dict) -> models.SurveyRecord:
    '''
    parses a survey record from a row in the dataset
    '''

    if row.get(field_mapping['student_id_field_name']) is None or len(row[field_mapping['student_id_field_name']].strip()) == 0:
        raise ValueError(f"found empty student id field in row: {row}")

    survey = models.SurveyRecord(parse_asurite(row[field_mapping['student_id_field_name']]))

    for field in field_mapping['preferred_students_field_names']:
        if re.search(r'\S', row[field]):
            survey.preferred_students.append(parse_asurite(row[field]).lower())

    survey.preferred_students = list(set(survey.preferred_students))

    for field in field_mapping['disliked_students_field_names']:
        if re.search(r'\S', row[field]):
            survey.disliked_students.append(
                parse_asurite(row[field]).lower())
    survey.disliked_students = list(set(survey.disliked_students))

    for field in field_mapping['availability_field_names']:
        avail_str = re.sub(r'\s', '', row[field].lower())
        survey.availability[field] = []
        if avail_str and re.search(r'\S', avail_str):
            survey.availability[field] = split_on_delimiters(
                avail_str, config.CONFIG_DATA["availability_values_delimiter"])

    if field_mapping.get('timezone_field_name'):
        survey.timezone = row[field_mapping['timezone_field_name']].strip()
    if (field_mapping.get("student_name_field_name") and row[field_mapping["student_name_field_name"]]):
        survey.student_name = row[field_mapping["student_name_field_name"]].strip()
    if (field_mapping.get("student_email_field_name") and row[field_mapping["student_email_field_name"]]):
        survey.student_email = row[field_mapping["student_email_field_name"]].strip()
    if (field_mapping.get("student_login_field_name") and row[field_mapping["student_login_field_name"]]):
        survey.student_login = row[field_mapping["student_login_field_name"]].strip()
    if (field_mapping.get("submission_timestamp_field_name") and row[field_mapping["submission_timestamp_field_name"]]):
        survey.submission_date = dt.datetime.strptime(
            row[field_mapping["submission_timestamp_field_name"]][:-4], '%Y/%m/%d %I:%M:%S %p')
    survey.provided_availability = has_availability(survey)

    return survey


def read_survey(field_mapping: models.SurveyFieldMapping, data_file_path: str) -> models.SurveyData:
    '''
    Loads the data from the survey.
    If there is a duplicate survey record, it will use (keep) the one with the submission
     date that is equal to or greater
    '''
    raw_rows: list[list[str]] = []
    records: list[models.SurveyRecord] = []
    with open(data_file_path, 'r', encoding='utf-8-sig') as data_file:
        raw_rows.extend(read_survey_raw(data_file))
        data_file.seek(0)
        records.extend(read_survey_records(field_mapping, data_file))

    return models.SurveyData(records, raw_rows)


def read_survey_from_io(field_mapping: models.SurveyFieldMapping, text_buffer: TextIOWrapper) -> models.SurveyData:
    '''
    Loads the survey data from an io buffer version of the data.
    If there is a duplicate survey record, it will use (keep) the one with the submission
     date that is equal to or greater
    '''
    raw_rows: list[list[str]] = []
    records: list[models.SurveyRecord] = []
    raw_rows.extend(read_survey_raw(text_buffer))
    text_buffer.seek(0)
    records.extend(read_survey_records(field_mapping, text_buffer))

    return models.SurveyData(records, raw_rows)


def read_survey_raw(data_file: TextIOWrapper) -> list[list[str]]:
    '''
    reads the csv into a 2d list.
    This is helpful for loading data that can be written elsewhere without changes
    '''
    rows = []
    reader = csv.reader(data_file)
    for row in reader:
        rows.append(row)

    return rows


def check_survey_field_headers(field_mapping: models.SurveyFieldMapping, fields):
    '''
    Checks that the fields in the survey data file match the field mapping
    '''
    valid_headers = True
    if field_mapping['student_id_field_name'] not in fields:
        __logger.error(__field_map_error_msg(field_mapping['student_id_field_name']))
        valid_headers = False
    for field in field_mapping['preferred_students_field_names']:
        if field not in fields:
            __logger.error(__field_map_error_msg(field))
            valid_headers = False
    for field in field_mapping['disliked_students_field_names']:
        if field not in fields:
            __logger.error(__field_map_error_msg(field))
            valid_headers = False
    for field in field_mapping['availability_field_names']:
        if field not in fields:
            __logger.error(__field_map_error_msg(field))
            valid_headers = False

    if field_mapping.get('submission_timestamp_field_name') and field_mapping['submission_timestamp_field_name'] not in fields:
        __logger.error(__field_map_error_msg(field_mapping['submission_timestamp_field_name']))
        valid_headers = False

    if valid_headers is False:
        raise ValueError("Headers from survey data file do not match the field mapping configuration")


def __field_map_error_msg(field: str):
    return f"Error: header '{field}' does not exist in the survey data file. Please check your field_mapping configuration"


def read_survey_records(field_mapping: models.SurveyFieldMapping, data_file: TextIOWrapper) -> list[models.SurveyRecord]:
    '''
    reads in a csv file using a csv dictreader and maps the fields back to the survey records
    '''
    reader = csv.DictReader(data_file)
    surveys: list[models.SurveyRecord] = []

    check_survey_field_headers(field_mapping, reader.fieldnames)

    for row in reader:
        survey = parse_survey_record(field_mapping, row)

        skip_user = False
        for idx, existing_survey_record in enumerate(surveys):
            if existing_survey_record.student_id == survey.student_id:
                if survey.submission_date >= existing_survey_record.submission_date:
                    print(
                        f'found duplicate record for id: {existing_survey_record.student_id}. Using record with timestamp: {survey.submission_date}')
                    surveys[idx] = survey
                else:
                    print(
                        f'skipped adding record with id: {existing_survey_record.student_id} and timestamp: {survey.submission_date}')
                skip_user = True

        if not skip_user:
            surveys.append(survey)

    preprocess_survey_data(surveys, field_mapping)

    return surveys


def read_groups(group_filename: str, survey_data: list[models.SurveyRecord]) -> list[models.GroupRecord]:
    '''
    Reads in grouping data into a list of GroupRecord objects
    '''
    with open(group_filename, 'r', encoding='utf-8-sig') as data_file:
        return read_groups_from_io(survey_data, data_file)


def read_groups_from_io(survey_data: list[models.SurveyRecord], text_buffer: TextIOWrapper) -> list[models.GroupRecord]:
    '''
    Loads/reads the grouping data into a list of Group Record objects from an io buffer
     version of the data.
    '''
    text_buffer.seek(0)
    groups: dict[str, models.GroupRecord] = {}
    is_header: bool = False
    reader = csv.DictReader(text_buffer)

    for row in reader:

        if is_header:
            is_header = False
            continue

        group_id = row["group id"]

        if not group_id in groups:
            groups[group_id] = models.GroupRecord(group_id, [])

        user = __get_user_by_id(row["student id"], survey_data)

        if not user is None:
            copied_user = copy.deepcopy(user)
            copied_user.group_id = group_id
            groups[group_id].members.append(copied_user)

    return list(groups.values())


def __get_user_by_id(student_id: str, survey_data: list[models.SurveyRecord]) -> Union[models.SurveyRecord, None]:
    '''
    Gets a SurveyRecord based on the student id
    '''

    for user in survey_data:
        if user.student_id == student_id:
            return user

    return None


def add_missing_students(survey: list[models.SurveyRecord], roster: list[str], avail_field: list[str]) -> list[models.SurveyRecord]:
    '''
    This method involves reading the survey data and adding any missing students from the student list
    to the survey data. This will be based on student id.
    '''
    new_survey_data = survey
    survey_students = list(map(lambda x: x.student_id, survey))

    # At this point, the survey_student list should have every student name from the survey data.
    # Next, we will check if all the students have answered the survey.

    missing_students = []
    for student in roster:
        if survey_students.count(student) == 0:
            missing_students.append(student)

    # At this point, all missing students should be in the missing students list.
    for asurite in missing_students:
        record = models.SurveyRecord(
            student_id=asurite,
        )
        # This code will use the function that adds availiability to all time slots.
        record.availability = wildcard_availability(avail_field)
        record.provided_survey_data = False
        record.provided_availability = False
        record.has_matching_availability = False
        new_survey_data.append(record)

    return new_survey_data


def read_roster(filename: str) -> list[str]:
    '''
    This method reads the roster file and returns the full roster of students
    '''
    roster = []

    with open(filename, 'r', encoding='utf-8-sig') as data_file:
        reader = csv.reader(data_file)
        next(reader)
        for row in reader:
            roster.append(row[0])

    return roster


def read_report_groups(report_filename: str, survey_data: list[models.SurveyRecord]) -> list[list[models.GroupRecord]]:
    '''
    Reads the specified groups from the individual tabs of an existing xlsx report file.
    '''

    group_sets: list[list[models.GroupRecord]] = []

    report_workbook = load_workbook(report_filename)

    for sheet_name in report_workbook.sheetnames:

        # Write the grouping data for each "individual" sheet/tab to an io buffer.
        if "individual" not in str.lower(sheet_name):
            continue

        text_buffer = StringIO()
        writer = csv.writer(text_buffer)

        # Variables for header information
        group_id_str: str = "group id"
        student_id_str: str = "student id"
        group_id_col: int = -1
        student_id_col: int = -1
        for row_idx, row in enumerate(report_workbook[sheet_name].rows):
            if row_idx == 0:
                # first row, write headers and determine group id and student id column numbers
                writer.writerow([group_id_str, student_id_str])
                for cell_idx, cell in enumerate(row):
                    if cell.value is None:
                        continue
                    if str.lower(cell.value) == str.lower(group_id_str):
                        group_id_col = cell_idx
                    elif str.lower(cell.value) == str.lower(student_id_str):
                        student_id_col = cell_idx
            elif group_id_col != -1 and student_id_col != -1:
                # write group information for each student
                writer.writerow(
                    [row[group_id_col].value, row[student_id_col].value])

        # load the group data from the io buffer
        text_buffer.seek(0)
        group_sets.append(read_groups_from_io(survey_data, text_buffer))

    return group_sets


def read_report_survey_data(report_filename: str, field_mappings: models.SurveyFieldMapping) -> models.SurveyData:
    '''
    Loads the survey data from the "survey_data" sheet (tab) of a previously generated xlsx report.
    '''

    # Write the data in the "survey_data" tab/sheet to an io buffer.

    report_workbook = load_workbook(report_filename)

    survey_data_sheet = report_workbook["survey_data"]

    text_buffer = StringIO()
    writer = csv.writer(text_buffer)
    for row in survey_data_sheet.rows:
        writer.writerow([cell.value for cell in row])

    # load and return the survey data from the io buffer
    text_buffer.seek(0)
    return read_survey_from_io(field_mappings, text_buffer)
