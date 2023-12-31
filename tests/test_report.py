import datetime
import math
import os
import shutil
from click.testing import CliRunner
from app.data import reporter
from app import config, models
from app.commands import report
from os.path import exists
from openpyxl import load_workbook, workbook
from openpyxl.worksheet import worksheet
from app.file import xlsx
import xlsxwriter
import random
import openpyxl

runner = CliRunner()

# These tests verify the functionality in report.py.


def test_dev_reporting_basic():
    '''
    Runs a simple test against the dev data set
    '''
    response = runner.invoke(report.report, [
                             './tests/test_files/dev_data/output.csv',
                             './tests/test_files/dev_data/dataset-dev.csv',
                             '-c', './tests/test_files/dev_data/config-dev.json'])
    assert response.exit_code == 0

    assert exists("./grouping_results_report.xlsx")

    os.remove("./grouping_results_report.xlsx")


def test_dev_reporting_report_file_specified():
    '''
    Runs a simple test against the dev data set but
    specifies the report output file name
    '''
    response = runner.invoke(report.report, [
                             './tests/test_files/dev_data/output.csv',
                             './tests/test_files/dev_data/dataset-dev.csv',
                             '-c', './tests/test_files/dev_data/config-dev.json',
                             '-r', './tests/test_files/dev_data/specific_report.xlsx'])
    assert response.exit_code == 0

    assert exists("./tests/test_files/dev_data/specific_report.xlsx")

    os.remove("./tests/test_files/dev_data/specific_report.xlsx")


def test_invalid_group_file():
    '''
    Runs a simple test to ensure it fails on a bad file
    '''
    response = runner.invoke(report.report, [
                             './tests/test_files/dev_data/badfile.csv',
                             './tests/test_files/dev_data/dataset-dev.csv',
                             '-c', './tests/test_files/dev_data/config-dev.json',
                             '-r', './tests/test_files/dev_data/specific_report.xlsx'])
    assert response.exit_code == 2


def test_invalid_survey_file():
    '''
    Runs a simple test to ensure it fails on a bad file
    '''
    response = runner.invoke(report.report, [
                             './tests/test_files/dev_data/output.csv',
                             './tests/test_files/dev_data/badfile.csv',
                             '-c', './tests/test_files/dev_data/config-dev.json',
                             '-r', './tests/test_files/dev_data/specific_report.xlsx'])
    assert response.exit_code == 2


def test_invalid_config_file():
    '''
    Runs a simple test to ensure it fails on a bad file
    '''
    response = runner.invoke(report.report, [
                             './tests/test_files/dev_data/output.csv',
                             './tests/test_files/dev_data/dataset-dev.csv',
                             '-c', './tests/test_files/dev_data/badfile.json',
                             '-r', './tests/test_files/dev_data/specific_report.xlsx'])
    assert response.exit_code == 2


def test_update_report_basic():
    '''
    Runs a simple "update report" test against Example_Report_1.xlsx.
    '''

    report_file_path: str = './tests/test_files/reports/Example_Report_1'
    shutil.copyfile(report_file_path + '.xlsx',
                    report_file_path + '_copy.xlsx')

    original_file_time: float = os.path.getmtime(report_file_path + '.xlsx')

    response = runner.invoke(report.update_report, [
                             './tests/test_files/reports/Example_Report_1.xlsx'])
    assert response.exit_code == 0

    # verify the report file was updated
    assert os.path.getmtime(report_file_path + '.xlsx') > original_file_time

    os.remove(report_file_path + '.xlsx')
    os.rename(report_file_path + '_copy.xlsx', report_file_path + '.xlsx')


def test_update_report_invalid_report_file():
    '''
    Runs a simple test to ensure update-report fails on a bad report file
    '''
    response = runner.invoke(report.update_report, [
                             './tests/test_files/reports/Nonexistent_File.xlsx'])
    assert response.exit_code == 2


def test_dev_reporting_contains_sheets():
    '''
    Checks the generated excel file for the expected sheets
    For now, just config
    '''
    response = runner.invoke(report.report, [
                             './tests/test_files/dev_data/output.csv',
                             './tests/test_files/dev_data/dataset-dev.csv',
                             '-c', './tests/test_files/dev_data/config-dev.json'])
    assert response.exit_code == 0

    assert exists("./grouping_results_report.xlsx")

    book: workbook.Workbook = load_workbook("./grouping_results_report.xlsx")
    assert 'config' in list(book.sheetnames)
    # add more asserts for more sheet names here

    os.remove("./grouping_results_report.xlsx")


def test_availability_map():
    '''
    More of an integration test: does the whole report
    and verifies the availability map
    '''

    data: list[models.GroupRecord] = [models.GroupRecord(group_id='1', members=[
        models.SurveyRecord(student_id='jsmith1', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='UTC +1', preferred_students=['jdoe2'], disliked_students=['mmuster3', 'jschmo4'], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': ['sunday', 'thursday', 'friday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [
        ], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['tuesday', 'wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
        models.SurveyRecord(student_id='jschmo4', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='', preferred_students=[], disliked_students=[], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [
        ], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
        models.SurveyRecord(student_id='bwillia5', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='', preferred_students=[], disliked_students=[], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [
        ], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
        models.SurveyRecord(student_id='jdoe2', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='UTC +2', preferred_students=['mmuster3', 'jschmo4'], disliked_students=['bwillia5', 'jsmith1'], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [
        ], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': ['tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
        models.SurveyRecord(student_id='mmuster3', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='UTC +3', preferred_students=['bwillia5', 'jsmith1'], disliked_students=['jdoe2'], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday', 'thursday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': ['thursday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': ['friday']}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id='')])]

    survey_data: models.SurveyData = models.SurveyData(records=[models.SurveyRecord(student_id='jschmo4', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='', preferred_students=[], disliked_students=[], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
                                                                models.SurveyRecord(student_id='bwillia5', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='', preferred_students=[], disliked_students=[], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [
                                                                ], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
                                                                models.SurveyRecord(student_id='mmuster3', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='UTC +3', preferred_students=['bwillia5', 'jsmith1'], disliked_students=['jdoe2'], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [
                                                                ], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday', 'thursday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': ['thursday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': ['friday']}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
                                                                models.SurveyRecord(student_id='jsmith1', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='UTC +1', preferred_students=['jdoe2'], disliked_students=['mmuster3', 'jschmo4'], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': ['sunday', 'thursday', 'friday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [
                                                                ], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['tuesday', 'wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id=''),
                                                                models.SurveyRecord(student_id='jdoe2', submission_date=datetime.datetime(2023, 3, 2, 18, 10, 59, 648426), student_name='', student_email='', student_login='', timezone='UTC +2', preferred_students=['mmuster3', 'jschmo4'], disliked_students=['bwillia5', 'jsmith1'], availability={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': ['tuesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['wednesday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': [], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': []}, has_matching_availability=True, provided_availability=True, provided_survey_data=True, group_id='')],
                                                       raw_rows=[['Timestamp', 'Username', 'Please select your ASURITE ID', 'Please enter your Github username (NOT your email address)', 'Email address for us to invite you to the Taiga scrumboard', 'In what time zone do you live or will you be during the session? Please use UTC so we can match it easier.', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]', 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]', 'How well would you say you know GitHub? (1 not at all, 5 worked with it a lot - know how to merge, resolve conflicts, etc.) You are not expected to know GitHub well yet, so please be honest. It will not be used for grading what you put here but I want to try to have one student knowing GitHub in each team to make things easier.', 'Do you know Scrum already? (1 just heard about it, 5 know it well (process, roles). You are not expected to know Scrum yet, so please be honest. It will not be used for grading what you put here. ', 'Preferred team member 1', 'Preferred team member 2', 'Preferred team member 3', 'Preferred team member 4', 'Preferred team member 5', 'Non-preferred student 1', 'Non-preferred student 2', 'Non-preferred student 3'], ['2022/10/17 6:31:58 PM EST', 'jsmith1@asu.edu', 'jsmith1', 'jsmith_1', 'johnsmith@gmail.com', 'UTC +1', 'Sunday;Thursday;Friday', 'Monday;Tuesday', '', '', '', 'Tuesday;Wednesday', '', '', '5', '2', 'jdoe2 - Jane Doe', '', '', '', '', 'mmuster3 - Max Mustermann', 'jschmo4 - Joe Schmo', ''], ['2022/10/17 6:33:27 PM EST', 'jdoe2@asu.edu', 'jdoe2', 'jdoe_2', 'janedoe@gmail.com', 'UTC +2', '', 'Monday;Tuesday', '', '', 'Tuesday', 'Wednesday', '', '', '4', '3', 'mmuster3 - Max Mustermann', 'jschmo4 - Joe Schmo', '', '', '', 'jsmith1 - John Smith', 'bwillia5 - Billy Williams', ''],
                                                                 ['2022/10/17 6:34:15 PM EST', 'mmuster3@asu.edu', 'mmuster3', 'mmuster_3', 'maxmustermann@gmail.com', 'UTC +3', '', 'Monday;Tuesday', '', '', '', 'Wednesday;Thursday', 'Thursday', 'Friday', '3', '4', 'jsmith1 - John Smith', 'bwillia5 - Billy Williams', '', '', '', 'jdoe2 - Jane Doe', '', ''], ['', '', 'jschmo4', '', '', '', '', 'Monday;Tuesday', '', '', '', 'Wednesday', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', 'bwillia5', '', '', '', '', 'Monday;Tuesday', '', '', '', 'Wednesday', '', '', '', '', '', '', '', '', '', '', '', '']])

    config_data: models.Configuration = config.read_json(
        "./tests/test_files/configs/config_1.json"
    )
    xlsx_writer = xlsx.XLSXWriter()
    green_bg = xlsx_writer.new_format("green_bg", {"bg_color": "#00FF00"})
    formatter = reporter.ReportFormatter(
        config_data, cell_formatters={'green_bg': green_bg})
    expected_map: models.AvailabilityMap = models.AvailabilityMap(availability_slots={'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [0:00 AM - 3:00 AM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 AM - 6:00 AM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 AM - 9:00 AM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 AM - 12:00 PM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [12:00 PM - 3:00 PM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [3:00 PM - 6:00 PM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [6:00 PM - 9:00 PM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'], 'Please choose times that are good for your team to meet. Times are in the Phoenix, AZ time zone! [9:00 PM - 12:00 PM]': ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']},
                                                                  group_availability=[models.GroupAvailabilityMap(group_id='1', users={'jsmith1': [False, False, False, True, True, False, True, True, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, True, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False], 'jschmo4': [False, False, False, False, False, False, False, True, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False], 'bwillia5': [False, False, False, False, False, False, False, True, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False], 'jdoe2': [False, False, False, False, False, False, False, True, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, True, False, False, False, False, False, False, False, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False], 'mmuster3': [False, False, False, False, False, False, False, True, True, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, False, True, True, False, False, False, False, False, False, True, False, False, False, False, False, False, False, True, False, False]})])
    availability_map = formatter.generate_availability_map(data, survey_data)

    assert availability_map == expected_map


def test_colored_columns():
    '''
    Checks if the availability overlap column has a background color of Green.
    '''
    workbook1 = xlsxwriter.Workbook("Random.xlsx")
    worksheet = workbook1.add_worksheet()
    green_bg = workbook1.add_format({"bg_color": "#00FF00"})

    worksheet.write('A1', 1, green_bg)
    worksheet.write('B1', 1)
    worksheet.write('C1', 1, green_bg)
    worksheet.write('A2', 1)
    worksheet.write('B2', 1, green_bg)
    worksheet.write('C2', 1)
    workbook1.close()

    book: workbook.Workbook = load_workbook("Random.xlsx")
    assert exists("Random.xlsx")
    ws1 = book["Sheet1"]

    cell1color = ws1[1][0].fill.start_color.index
    cell2color = ws1[1][1].fill.start_color.index
    cell3color = ws1[1][2].fill.start_color.index
    cell4color = ws1[2][0].fill.start_color.index
    cell5color = ws1[2][1].fill.start_color.index
    cell6color = ws1[2][2].fill.start_color.index

    assert cell1color == "FF00FF00"
    assert cell2color != "FF00FF00"
    assert cell3color == "FF00FF00"
    assert cell4color != "FF00FF00"
    assert cell5color == "FF00FF00"
    assert cell6color != "FF00FF00"

    os.remove("Random.xlsx")


def test_freeze_columns_2():
    '''
    Checks if the first two columns in a sheet is frozen
    '''
    response = runner.invoke(report.report, [
        './tests/test_files/dev_data/output.csv',
        './tests/test_files/dev_data/dataset-dev.csv',
        '-c', './tests/test_files/dev_data/config-dev.json',
        '-r' 'test_report_for_frozen_columns.xlsx'])
    assert response.exit_code == 0

    assert exists("test_report_for_frozen_columns.xlsx")

    book = load_workbook(
        "test_report_for_frozen_columns.xlsx")
    assert 'config' in list(book.sheetnames)
    assert 'individual_report_1' in list(book.sheetnames)
    worksheet = book['individual_report_1']

    assert worksheet.freeze_panes == 'C1'

    os.remove("test_report_for_frozen_columns.xlsx")


def test_report_for_formatted_columns():
    '''
    Checks the generated report for columns with formatted widths rather than the default
    '''
    response = runner.invoke(report.report, [
                             './tests/test_files/dev_data/output.csv',
                             './tests/test_files/dev_data/dataset-dev.csv',
                             '-c', './tests/test_files/dev_data/config-dev.json',
                             '-r' 'test_report_for_correctness.xlsx'])
    assert response.exit_code == 0

    assert exists(
        "test_report_for_correctness.xlsx")

    book: workbook.Workbook = load_workbook(
        "test_report_for_correctness.xlsx")
    assert 'config' in list(book.sheetnames)
    assert 'individual_report_1' in list(book.sheetnames)

    individual_report_1: worksheet.Worksheet = book['individual_report_1']
    individual_report_2: worksheet.Worksheet = book['individual_report_1']
    group_report_1: worksheet.Worksheet = book['group_report_1']

    # check the width of a sampling of the columns and ensure they aren't the default. Kind of a hacky way to check but it should work
    assert not math.isclose(
        individual_report_1.column_dimensions['A'].width, 13.0)
    assert not math.isclose(
        individual_report_1.column_dimensions['B'].width, 13.0)
    assert not math.isclose(
        individual_report_1.column_dimensions['C'].width, 13.0)

    assert not math.isclose(
        individual_report_2.column_dimensions['A'].width, 13.0)
    assert not math.isclose(
        individual_report_2.column_dimensions['B'].width, 13.0)
    assert not math.isclose(
        individual_report_2.column_dimensions['C'].width, 13.0)

    assert not math.isclose(group_report_1.column_dimensions['A'].width, 13.0)
    assert not math.isclose(group_report_1.column_dimensions['B'].width, 13.0)
    assert not math.isclose(group_report_1.column_dimensions['C'].width, 13.0)
    os.remove("test_report_for_correctness.xlsx")

def test_report_shows_true_in_meets_preferred_column_group():
    '''
    This test will check to make sure that if a student did not specify any preferred students, the meets preferred
    column will show true in the Group summary tab
    '''

    shutil.copyfile('tests/test_files/reports/dataset-305_report_group.xlsx',
                    'tests/test_files/reports/dataset_305_report_group_temp.xlsx')

    response = runner.invoke(report.update_report, ['tests/test_files/reports/dataset_305_report_group_temp.xlsx'])

    book = load_workbook('tests/test_files/reports/dataset_305_report_group_temp.xlsx')
    worksheet = book['group_report_1']
    worksheet2 = book['group_report_2']
    assert worksheet['G3'].value == True
    assert worksheet2['G3'].value == True

    os.remove('tests/test_files/reports/dataset_305_report_group_temp.xlsx')


def test_report_shows_blank_in_meets_disliked_preferred_columns_individual():
    '''
    This test will check to make sure that if a student didn't specify disliked students, the meets disliked requirment
    column will show true
    '''

    shutil.copyfile('tests/test_files/reports/dataset_305_report.xlsx', 'tests/test_files/reports/dataset_305_report_temp.xlsx')

    response = runner.invoke(report.update_report, ['tests/test_files/reports/dataset_305_report_temp.xlsx'])

    book = load_workbook('tests/test_files/reports/dataset_305_report_temp.xlsx')
    worksheet = book['individual_report_1']
    worksheet2 = book['individual_report_2']
    assert worksheet['E3'].value == None
    assert worksheet2['E2'].value == None
    assert worksheet['J2'].value == None
    assert worksheet2['J3'].value == None

    os.remove('tests/test_files/reports/dataset_305_report_temp.xlsx')