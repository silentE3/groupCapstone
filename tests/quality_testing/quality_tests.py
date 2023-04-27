'''
These tests were created for the quality plan created for the project.  They will not run every time there is a pull
request.  Some of the quality tests were alredy part of our testing regime, and those will remain in the tests package.
'''

import math
import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from click.testing import CliRunner
from tests.test_utils.helper_functions import verify_all_students
from tests.test_utils.helper_functions import verify_groups
from app import config, models
from app.commands import group, report as report_imp
from app.data import load


runner = CliRunner()

def test_group_sizing_t01():
    '''
    Test of grouping 12 students with a target group size of 4 (divides evenly), without any +/- margin. 

    The expected outcome is that the tool provides solutions consisting of 3 groups of 4 students.
    '''

    response = runner.invoke(group.group, [
        './tests/test_files/survey_results/test_group_1.csv', '--configfile', './tests/test_files/configs/test_group_1_config.json'])
    assert response.exit_code == 0

    expected_min_num_groups = 3
    expected_max_num_groups = 3
    expected_students = ['jsmith1', 'jdoe2',
                         'mmuster3', 'jschmo4', 'bwillia5', 'mbrown6', 'charles7', 'carl8', 'elee9', 'cred10', 'bobbylee11', 'jrogan12']

    verify_groups('./tests/test_files/survey_results/test_group_1_report.xlsx', expected_min_num_groups,
                  expected_max_num_groups, expected_students)

    # Verify "Error:" is NOT included in the output
    assert "Error:" not in response.output

    os.remove(
        './tests/test_files/survey_results/test_group_1_report.xlsx')


def test_group_sizing_t02():
    '''
    Test of grouping 16 students with a target group size of 3 (does not divide evenly,
    but still possible to maintain +1, [group sizes of 3 or 4]).
    
    '''

    response = runner.invoke(group.group, [
                             './tests/test_files/survey_results/Example_Survey_Results_16.csv',
                             '--configfile', './tests/test_files/configs/config_16.json',
                             '--reportfile', './tests/test_files/survey_results/test_16_report.xlsx'])
    assert response.exit_code == 0

    expected_students = ['adumble4', 'triddle8', 'dmalfoy7',
                         'rweasle3', 'hgrange2', 'rhagrid5', 'hpotter1', 'nlongbo6',
                         'adumble4_2', 'triddle8_2', 'dmalfoy7_2',
                         'rweasle3_2', 'hgrange2_2', 'rhagrid5_2', 'hpotter1_2', 'nlongbo6_2']
    verify_groups('./tests/test_files/survey_results/test_16_report.xlsx', 4,
                  5, expected_students)
    # Verify "Error:" is NOT included in the output
    assert "Error:" not in response.output

    os.remove('./tests/test_files/survey_results/test_16_report.xlsx')


def test_group_sizing_t03():
    '''
    Test of grouping 19 students with a target group size of 5 and a tolerance of -1.
    [group sizes of 4 or 5]
    '''
    response = runner.invoke(group.group, [
                             './tests/test_files/survey_results/Example_Survey_Results_19.csv',
                             '--configfile', './tests/test_files/configs/config_19.json',
                             '--reportfile', './tests/test_files/survey_results/test_19_report.xlsx'])
    assert response.exit_code == 0

    expected_students = ['uenterprise2', 'uhornet3', 'uyorktown1',
                         'ulexington4', 'usaratoga5', 'jakagi6',
                         'jkaga7', 'jzuikaku8', 'jshokaku9',
                         'uenterprise2_2', 'uhornet3_2', 'uyorktown1_2',
                         'ulexington4_2', 'usaratoga5_2', 'jakagi6_2',
                         'jkaga7_2', 'jzuikaku8_2', 'jshokaku9_2', 'jhiryu10']
    verify_groups('./tests/test_files/survey_results/test_19_report.xlsx', 3,
                  4, expected_students)
    # Verify "Error:" is NOT included in the output
    assert "Error:" not in response.output

    os.remove('./tests/test_files/survey_results/test_19_report.xlsx')


def test_group_sizing_t04():
    '''
    Test of grouping 18 students with a target group size of 5 +/- 1 (testing +/-
     1 variability here).
    The expected outcome is that the tool provides solutions consisting
     of groups whose sizes all fall within the range [4, 6].
    '''

    response = runner.invoke(group.group, [
                             './tests/test_files/survey_results/Example_Survey_Results_18.csv',
                             '--configfile', './tests/test_files/configs/config_18.json',
                             '--reportfile', './tests/test_files/survey_results/test_18_report.xlsx'])
    assert response.exit_code == 0

    expected_students = ['adumble4', 'triddle8', 'dmalfoy7',
                         'rweasle3', 'hgrange2', 'rhagrid5', 'hpotter1', 'nlongbo6',
                         'adumble4_2', 'triddle8_2', 'dmalfoy7_2',
                         'rweasle3_2', 'hgrange2_2', 'rhagrid5_2', 'hpotter1_2', 'nlongbo6_2',
                         'hpotter1_3', 'nlongbo6_3']
    verify_groups('./tests/test_files/survey_results/test_18_report.xlsx', 4,
                  6, expected_students)
    # Verify "Error:" is NOT included in the output
    assert "Error:" not in response.output

    os.remove('./tests/test_files/survey_results/test_18_report.xlsx')


def test_group_sizing_invalid_t05():
    '''
    Test of grouping 10 students with a target group size of 7 and a tolerance of +/-1.

    The expected outcome is an error since it is not possible to adhere to these group
    sizing constraints.
    '''
    response = runner.invoke(group.group, [
                             './tests/test_files/survey_results/Example_Survey_Results_10.csv', '--configfile', './tests/test_files/configs/config_10.json', '--reportfile', './tests/test_files/survey_results/test_10_report.xlsx'])
    assert response.exit_code == 0

    # Verify "Error:" IS included in the output
    assert "Error:" in response.output
    assert not os.path.exists('./tests/test_files/survey_results/test_10_report.xlsx')


def test_complete_solution_t06():
    '''
    This test is to verify that when a surveyfile containing 100 students is provided to the program, with a target
    group size of 5 with a tolarance of +1, the program will produce a grouping solution with all 100 students grouped
    '''
    complete_solution = runner.invoke(group.group, ['./tests/test_files/survey_results/Example_Survey_Results_100.csv',
                                                    '-c', './tests/test_files/configs/config_100.json'])
    expected_min_num_groups = 17
    expected_max_num_groups = 20
    expected_students = list(f'asurite{x}' for x in range(100))

    assert complete_solution.exit_code == 0

    verify_all_students('./tests/test_files/survey_results/Example_Survey_Results_100_report.xlsx', expected_min_num_groups,
                        expected_max_num_groups, expected_students)

    os.remove('./tests/test_files/survey_results/Example_Survey_Results_100_report.xlsx')


def test_deterministic_scoring_t07():
    '''
    This test verifies that solution scores generated by the tool are deterministic
    under constant grouping conditions.
    '''
    ID_COL_NUM: int = 0
    GROUP_COL_NUM: int = 1
    SWAP_ROW_NUM_1: int = 2
    SWAP_ROW_NUM_2: int = 7
    FIRST_STUD_IDX = 0
    SECOND_STUD_IDX = 1

    def __get_scores(report_wb: Workbook) -> list[float]:
        scores: list[float] = []
        overall_sheets = [report_wb["overall_report_1"], report_wb["overall_report_2"]]
        for overall_sheet in overall_sheets:
            for col_num, cell in enumerate(overall_sheet[1]):  # first row
                if str.strip(cell.value) == "Score":
                    scores.append(float(overall_sheet[2][col_num].value))  # second row value
        return scores

    # Note, swap info is saved as:
    # ((first_student_id, first_student_orig_group), (second_student_id, second_student_orig_group))
    def __move_swapped_students(report_wb: Workbook, swap_info: list[tuple[tuple, tuple]], to_orig_group: bool):
        individual_sheets: list[Worksheet] = [report_wb["individual_report_1"], report_wb["individual_report_2"]]
        for sheet_idx, individual_sheet in enumerate(individual_sheets):
            for row_idx, stud_id_cell in enumerate(individual_sheet['A']):
                row_num = row_idx + 1
                if stud_id_cell.value == swap_info[sheet_idx][FIRST_STUD_IDX][ID_COL_NUM]:  # first swapped student id
                    target_student_idx = FIRST_STUD_IDX if to_orig_group else SECOND_STUD_IDX
                    individual_sheet[row_num][GROUP_COL_NUM].value = swap_info[sheet_idx][target_student_idx][GROUP_COL_NUM]
                # second swapped student id
                elif stud_id_cell.value == swap_info[sheet_idx][SECOND_STUD_IDX][ID_COL_NUM]:
                    target_student_idx = SECOND_STUD_IDX if to_orig_group else FIRST_STUD_IDX
                    individual_sheet[row_num][GROUP_COL_NUM].value = swap_info[sheet_idx][target_student_idx][GROUP_COL_NUM]
        report_wb.save('tests/test_files/survey_results/dataset-35_students_report.xlsx')
        report_wb.close()

    def __compare_scores(expected_scores: list[float], actual_scores: list[float]):
        for idx, expected_score in enumerate(expected_scores):  # first row
            assert math.isclose(expected_score, actual_scores[idx], rel_tol=0.0001)

    # An input of 35 students is provided to the program for grouping, along with
    # valid group sizing constraints (the specifics of which are irrelevant).
    runner.invoke(group.group, ['./tests/test_files/survey_results/dataset-35_students.csv',
                                '-c', './tests/test_files/configs/config-35_students.json'])
    assert os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    if not os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx'):
        return

    report: Workbook = load_workbook('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    # Save the scores from the grouping process as the initial scores
    initial_scores: list[float] = __get_scores(report)

    # Within the report, the group assignments for two students are swapped within each solution
    individual_sheets = [report["individual_report_1"], report["individual_report_2"]]
    # Note, swaps will be saved as:
    # ((first_student_id, first_student_orig_group), (second_student_id, second_student_orig_group))
    swapped_students: list[tuple[tuple, tuple]] = []
    for idx, individual_sheet in enumerate(individual_sheets):
        # save the swap information
        swapped_students.append(((individual_sheet[SWAP_ROW_NUM_1][ID_COL_NUM].value, individual_sheet[SWAP_ROW_NUM_1][GROUP_COL_NUM].value), (
            individual_sheet[SWAP_ROW_NUM_2][ID_COL_NUM].value, individual_sheet[SWAP_ROW_NUM_2][GROUP_COL_NUM].value)))
        # swap the student's group assignments
        individual_sheet[SWAP_ROW_NUM_2][GROUP_COL_NUM].value = swapped_students[idx][FIRST_STUD_IDX][GROUP_COL_NUM]
        individual_sheet[SWAP_ROW_NUM_1][GROUP_COL_NUM].value = swapped_students[idx][SECOND_STUD_IDX][GROUP_COL_NUM]
    report.save('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    report.close()

    # Process the updated report to get the new scores
    runner.invoke(report_imp.update_report, ['./tests/test_files/survey_results/dataset-35_students_report.xlsx'])
    assert os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    if not os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx'):
        return
    report = load_workbook('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    # Save the swapped scores
    swapped_scores = __get_scores(report)

    # Within the report, the students are returned to their initial group assignments
    __move_swapped_students(report, swapped_students, True)

    # Process the updated report to get the new scores
    runner.invoke(report_imp.update_report, ['./tests/test_files/survey_results/dataset-35_students_report.xlsx'])
    assert os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    if not os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx'):
        return
    report = load_workbook('tests/test_files/survey_results/dataset-35_students_report.xlsx')

    # The scores for the solutions output by the program are compared against the ‘initial scores’. The
    # expected result is that the error between these values is less than 0.01%
    __compare_scores(initial_scores, __get_scores(report))

    # Within the report, the students are returned to their swapped group assignments
    __move_swapped_students(report, swapped_students, False)

    # Process the updated report to get the new scores
    runner.invoke(report_imp.update_report, ['./tests/test_files/survey_results/dataset-35_students_report.xlsx'])
    assert os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    if not os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx'):
        return
    report = load_workbook('tests/test_files/survey_results/dataset-35_students_report.xlsx')

    # The scores for the solutions output by the program are compared against the ‘swapped scores’. The
    # expected result is that the error between these values is less than 0.01%
    __compare_scores(swapped_scores, __get_scores(report))

    os.remove('tests/test_files/survey_results/dataset-35_students_report.xlsx')


def test_input_file_t09():
    '''
    This test verifies that the grouping process uses the specified input file.
    '''

    # invoke the group command with a specified survey file that is valid
    runner.invoke(group.group, ['./tests/test_files/survey_results/dataset-35_students.csv',
                                '-c', './tests/test_files/configs/config-35_students.json'])
    assert os.path.exists('tests/test_files/survey_results/dataset-35_students_report.xlsx')
    if not os.path.exists:
        return
    config_data = config.read_json('./tests/test_files/configs/config-35_students.json')

    # read the "raw" (csv) survey data
    survey_data_raw = load.read_survey(config_data['field_mappings'],
                                       './tests/test_files/survey_results/dataset-35_students.csv')

    # read the survey data from the xlsx file that was output by the group command
    survey_data_from_report = load.read_report_survey_data(
        'tests/test_files/survey_results/dataset-35_students_report.xlsx', config_data['field_mappings'])

    # Verify that the raw data read in matches what was read in from the report file
    assert survey_data_raw == survey_data_from_report

    os.remove('tests/test_files/survey_results/dataset-35_students_report.xlsx')


def test_report_filename_t10():
    '''
    This test invokes the group command and verifies that it is writing to a .xlsx file of the name that is expected
    according to our code that determines the name of the report file based on the name of the survey file provided at
    input. To pass, the group command must create the output file with the correct name, and inform the user with a
    message.
    '''

    response = runner.invoke(group.group, ['tests/test_files/survey_results/Example_Survey_Results_1.csv',
                                           '--configfile', 'tests/test_files/configs/config_1.json'])
    assert response.exit_code == 0
    assert os.path.exists('tests/test_files/survey_results/Example_Survey_Results_1_report.xlsx')
    assert "Writing report to: tests/test_files/survey_results/Example_Survey_Results_1_report.xlsx" in response.output
    os.remove('tests/test_files/survey_results/Example_Survey_Results_1_report.xlsx')


def test_group_empty_survey_t11():
    '''
    This test passes a survey file to the group command that has no students in it. It passes if there are no groups in
    the report, and there is a message displayed to the user telling them there was an error with the survey supplied.
    '''

    response = runner.invoke(group.group, ['tests/test_files/survey_results/Example_Survey_Results_3.csv',
                                           '--configfile', 'tests/test_files/configs/config_3.json'])
    assert response.exit_code == 0
    assert "Writing report to: tests/test_files/survey_results/Example_Survey_Results_3_report.xlsx" not in response.output
    assert "Error: No student surveys found in the input datafile." in response.output


def test_read_survey_raw_quality_t12():
    """
    tests that loading the raw records of a survey file reads the right number of rows
    and the first row is the header
    """
    rows = []
    with open('./tests/test_files/survey_results/Example_Survey_Results_1_full.csv', 'r', encoding="UTF-8") as file:
        rows.extend(load.read_survey_raw(file))
    assert len(rows) == 4
    assert rows[0][0] == 'Timestamp'

    assert rows[1][0] == '2022/10/17 6:31:58 PM EST'
    assert rows[2][0] == '2022/10/17 6:31:58 PM EST'
    assert rows[3][0] == "2022/10/17 6:31:58 PM EST"

    assert rows[1][1] == 'jsmith1@asu.edu'
    assert rows[2][1] == 'jdoe2@asu.edu'
    assert rows[3][1] == 'mmuster3@asu.edu'

    assert rows[1][2] == 'jsmith1'
    assert rows[2][2] == 'jdoe2'
    assert rows[3][2] == 'mmuster3'

    assert rows[1][3] == 'John Smith'
    assert rows[2][3] == 'Jane Doe'
    assert rows[3][3] == 'Max Munster'

    assert rows[1][4] == 'jsmith_1'
    assert rows[2][4] == 'jdoe_2'
    assert rows[3][4] == 'mmuster_3'

    assert rows[1][5] == 'johnsmith@gmail.com'
    assert rows[2][5] == 'janedoe@gmail.com'
    assert rows[3][5] == 'maxmustermann@gmail.com'

    assert rows[1][6] == 'UTC +1'
    assert rows[2][6] == 'UTC +2'
    assert rows[3][6] == 'UTC +3'

    assert rows[1][7] == 'Sunday;Thursday;Friday'
    assert rows[2][7] == ''
    assert rows[3][7] == ''

    assert rows[1][8] == 'Monday'
    assert rows[2][8] == ''
    assert rows[3][8] == ''

    assert rows[1][9] == ''
    assert rows[2][9] == ''
    assert rows[3][9] == ''

    assert rows[1][10] == ''
    assert rows[2][10] == ''
    assert rows[3][10] == ''

    assert rows[1][11] == ''
    assert rows[2][11] == 'Tuesday'
    assert rows[3][11] == ''

    assert rows[1][12] == ''
    assert rows[2][12] == 'Wednesday'
    assert rows[3][12] == ''

    assert rows[1][13] == ''
    assert rows[2][13] == ''
    assert rows[3][13] == 'Thursday'

    assert rows[1][14] == ''
    assert rows[2][14] == ''
    assert rows[3][14] == 'Friday'

    assert rows[1][15] == '5'
    assert rows[2][15] == '4'
    assert rows[3][15] == '3'

    assert rows[1][16] == '2'
    assert rows[2][16] == '3'
    assert rows[3][16] == '4'

    assert rows[1][17] == 'jdoe2 - Jane Doe'
    assert rows[2][17] == 'mmuster3 - Max Mustermann'
    assert rows[3][17] == 'jsmith1 - John Smith'

    assert rows[1][18] == ''
    assert rows[2][18] == 'jschmo4 - Joe Schmo'
    assert rows[3][18] == 'bwillia5 - Billy Williams'

    assert rows[1][19] == ''
    assert rows[2][19] == ''
    assert rows[3][19] == ''

    assert rows[1][20] == ''
    assert rows[2][20] == ''
    assert rows[3][20] == ''

    assert rows[1][21] == ''
    assert rows[2][21] == ''
    assert rows[3][21] == ''

    assert rows[1][22] == 'mmuster3 - Max Mustermann'
    assert rows[2][22] == 'jsmith1 - John Smith'
    assert rows[3][22] == 'jdoe2 - Jane Doe'

    assert rows[1][23] == 'jschmo4 - Joe Schmo'
    assert rows[2][23] == 'bwillia5 - Billy Williams'
    assert rows[3][23] == ''

    assert rows[1][24] == ''
    assert rows[2][24] == ''
    assert rows[3][24] == ''


def test_read_survey_wrong_file_type_quality_t13():
    """
    tests that loading a survey file that isn't the right type of file will raise an exception
    """

    response = runner.invoke(group.group, [
        './tests/test_files/reports/Example_Report_1.xlsx',
        '--configfile',
        './tests/test_files/configs/config_19.json',
        '--reportfile',
        './tests/test_files/survey_results/test_19_report.xlsx'])

    assert response.exit_code != 0


def test_read_roster_t14():
    '''
    This test verifies that every student listed in the roster is included in the
    grouping process, regardless of whether they are included in the survey data.
    '''

    # NOTE: The survey data only includes asurites 1 through 5, whereas the
    # roster (allstudentsfile) includes asurites 1 through 8.
    response = runner.invoke(group.group, [
                             './tests/test_files/survey_results/Example_Survey_Results_T14.csv',
                             '--configfile', './tests/test_files/configs/config_T14.json',
                             '--allstudentsfile', './tests/test_files/example_roster_3.csv'])

    assert response.exit_code == 0

    expected_students = ['asurite1',
                         'asurite2',
                         'asurite3',
                         'asurite4',
                         'asurite5',
                         'asurite6',
                         'asurite7',
                         'asurite8']

    # Verify all of the expected students have been grouped
    report_wb: Workbook = load_workbook('tests/test_files/survey_results/Example_Survey_Results_T14_report.xlsx')
    individual_sheets = [report_wb["individual_report_1"], report_wb["individual_report_2"]]
    for individual_sheet in individual_sheets:
        row_num: int = 2
        for col_num, cell in enumerate(individual_sheet[1]):  # first row
            if cell.value is None:
                continue
            if str.strip(cell.value) == "Student Id":
                while (individual_sheet[row_num][col_num].value) is not None:  # second row value
                    assert individual_sheet[row_num][col_num].value in expected_students
                    row_num += 1
        assert row_num-2 == len(expected_students)  # minus 2 b/c started at 2

    # Verify "Error:" is NOT included in the output
    assert "Error:" not in response.output

    os.remove('./tests/test_files/survey_results/Example_Survey_Results_T14_report.xlsx')


def test_load_missing_students_t15():
    '''
    Tests the add missing student function with 2 missing students.
    '''
    result = []

    ids = []
    ids.append("asurite1")
    ids.append("asurite2")
    ids.append("asurite3")
    ids.append("asurite4")
    ids.append("asurite5")
    ids.append("asurite6")

    survey = []
    student1 = models.SurveyRecord("asurite1")
    student2 = models.SurveyRecord("asurite2")
    student3 = models.SurveyRecord("asurite3")
    student4 = models.SurveyRecord("asurite4")
    survey.append(student1)
    survey.append(student2)
    survey.append(student3)
    survey.append(student4)

    fields = [
        "0 to 3 AM",
        "3 to 6 AM",
        "6 to 9 AM",
        "9 to 12 PM",
        "12 to 3 PM",
        "3 to 6 PM",
        "6 to 9 PM",
        "9 to 12 AM"
    ]

    result = load.add_missing_students(survey, ids, fields)

    assert len(result) == 6
    assert result[4].availability == {
        "0 to 3 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "3 to 6 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "6 to 9 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "9 to 12 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "12 to 3 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "3 to 6 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "6 to 9 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "9 to 12 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    }
    assert result[5].availability == {
        "0 to 3 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "3 to 6 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "6 to 9 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "9 to 12 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "12 to 3 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "3 to 6 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "6 to 9 PM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'],
        "9 to 12 AM": ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
    }
    assert not result[4].provided_survey_data
    assert not result[5].provided_survey_data


def test_config_field_names_t16():
    '''
    T-16: Tests the condition where a subset of field names in the data do not exist in the config. 
    This should ignore any fields that are not defined in the config file.
    '''
    conf = config.read_json('./tests/test_files/configs/test_16_config.json')

    survey = load.read_survey(conf['field_mappings'], './tests/test_files/survey_results/quality_test_17.csv')

    # check the survey to see if the fields that are not defined in the config file are ignored
    for record in survey.records:
        assert list(record.availability.keys()) == conf['field_mappings']['availability_field_names']
        assert len(record.disliked_students) <= len(conf['field_mappings']['disliked_students_field_names'])
        assert len(record.preferred_students) <= len(conf['field_mappings']['preferred_students_field_names'])


def test_config_field_names_t17():
    '''
    T-17: Test the condition where field names in config do not exist in the data
    '''
    response = runner.invoke(group.group, [
                             './tests/test_files/survey_results/quality_test_17.csv',
                             '--configfile', './tests/test_files/configs/invalid_config.json',
                             '--reportfile', './tests/test_files/survey_results/test_17_report.xlsx'])
    assert response.exit_code != 0

    # Verify "Error:" is included in the output
    assert "Error" in response.output
